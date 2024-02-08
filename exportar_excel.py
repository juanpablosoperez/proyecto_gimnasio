from tkinter import Tk, Label, Frame, Button, StringVar, filedialog, messagebox
import time
import pandas as pd
import tkinter as tk
import sqlite3
import openpyxl
from tkinter.ttk import Combobox  # Asegúrate de importar Combobox desde tkinter.ttk
from datetime import datetime

# Establecer la configuración del escritor de Excel
pd.set_option('io.excel.xlsx.writer', 'xlsxwriter')


class VentanaExportar:
    def __init__(self, ventana_exportar):
        self.window = ventana_exportar
        self.window.title("Exportar Excel")
        self.window.geometry("980x520")

        # Obtener las dimensiones de la pantalla
        screen_width = ventana_exportar.winfo_screenwidth()
        screen_height = ventana_exportar.winfo_screenheight()

        # Calcular las coordenadas para centrar la ventana
        x = (screen_width - 980) // 2
        y = (screen_height - 520) // 2

        # Establecer la geometría de la ventana
        ventana_exportar.geometry(f"980x520+{x}+{y}")

        self.window.resizable(width=False, height=False)
        self.window.config(bg="#363636", bd=10)

        # Crear el marco principal
        marco_principal = Frame(self.window, bg="#363636")
        marco_principal.pack(side="top", fill="x", pady=10, padx=10, anchor="nw")

        # Botón para regresar a la pantalla principal
        estilo_boton_ppal = {
            "bd": 0,
            "fg": "white",
            "relief": "flat",
            "font": ("Arial", 11),
            "padx": 10,
            "pady": 5,
            "bg": "#363636"
        }

        boton_principal = Button(
            marco_principal,
            text="Pantalla Principal",
            command=self.volver_pantalla_principal,
            **estilo_boton_ppal
        )
        boton_principal.pack(side="left")

        marco_fecha_hora = Frame(self.window, bg="#212121", bd=2, relief="solid")
        marco_fecha_hora.place(relx=1, rely=1, anchor="se", x=-10, y=-10)

        # Etiqueta para la fecha y hora
        self.etiqueta_fecha_hora = Label(marco_fecha_hora, text="", font=("Arial", 12), bg="#212121", fg="#7FFFD4")
        self.etiqueta_fecha_hora.pack(padx=10, pady=5)

        # Agregar el título en negrita centrado sobre el marco de exportar
        titulo_label = Label(self.window, text="EXPORTAR EXCEL", font=("Arial", 24, "bold"), bg="#363636",
                             fg="#7FFFD4")
        titulo_label.pack(pady=(0, 10))

        # Agregar un nuevo marco para exportar datos
        marco_exportar = Frame(self.window, bg="#363636", bd=4, relief="solid", highlightbackground="#7FFFD4",
                               highlightthickness=0.5)
        marco_exportar.place(relx=0.5, rely=0.5, anchor="center", relwidth=0.5, relheight=0.4)

        # Agregar etiqueta descriptiva
        etiqueta_descriptiva = Label(marco_exportar, text="Seleccionar actividad para exportar datos a Excel",
                                     font=("Arial", 14), bg="#363636", fg="#7FFFD4")
        etiqueta_descriptiva.pack(pady=10)

        # Crear el marco para los botones de Excel y Gráficos
        marco_botones = Frame(marco_exportar, bg="#363636")
        marco_botones.pack(pady=20)

        # Obtener actividades disponibles desde la base de datos
        conn = sqlite3.connect("gimnasio.db")
        c = conn.cursor()
        c.execute("SELECT nombreActividad FROM Actividad")
        actividades_db = c.fetchall()
        conn.close()

        # Asociar las actividades al Combobox de selección de actividad
        self.actividades = [nombre_actividad for nombre_actividad, in actividades_db]
        self.actividad_seleccionada = StringVar(marco_botones)
        self.actividad_seleccionada.set("--SELECCIONAR ACTIVIDAD--")
        self.combobox_actividad = Combobox(marco_botones, textvariable=self.actividad_seleccionada,
                                           values=["--SELECCIONAR ACTIVIDAD--"] + self.actividades,
                                           state='readonly',width=27)
        self.combobox_actividad.grid(row=0, column=0, padx=(15, 20), pady=(0, 15))

        # Crear el botón "Exportar"
        estilo_boton_exportar = {
            "bd": 0,
            "fg": "white",
            "relief": "flat",
            "font": ("Arial", 12),
            "padx": 20,
            "pady": 10,
            "bg": "#4CAF50"
        }
        self.boton_exportar = Button(
            marco_botones,
            text="Exportar",
            command=self.exportar_a_excel,
            **estilo_boton_exportar
        )
        self.boton_exportar.grid(row=1, column=0, pady=(0, 10))
        # Deshabilitar el botón Exportar al inicio
        self.boton_exportar["state"] = "disabled"

        # Actualizar la fecha y hora cada segundo
        self.actualizar_fecha_hora()

    def exportar_a_excel(self):
        # Verificar si se ha seleccionado una actividad
        if self.actividad_seleccionada.get() == "--SELECCIONAR ACTIVIDAD--":
            messagebox.showwarning("Advertencia", "Por favor, selecciona una actividad antes de exportar.")
            return

        # Obtener la actividad seleccionada
        actividad_seleccionada = self.actividad_seleccionada.get()

        # Paso 1: Conexión a la base de datos
        conn = sqlite3.connect("gimnasio.db")
        c = conn.cursor()

        # Paso 2: Verificar la consulta SQL
        print("Actividad seleccionada:", actividad_seleccionada)

        consulta_sql = """
            SELECT c.nombre, c.apellido, p.fechaPago, SUM(p.montoTotal) AS MontoTotal
            FROM Pagos p
            JOIN Clientes c ON p.idCliente = c.idCliente
            JOIN Actividad a ON p.idActividad = a.idActividad
            WHERE a.nombreActividad = ?
            GROUP BY c.nombre, c.apellido, p.fechaPago
            ORDER BY p.fechaPago DESC
        """
        print("Consulta SQL:", consulta_sql)

        c.execute(consulta_sql, (actividad_seleccionada,))
        resultados = c.fetchall()

        # Paso 3: Verificar los resultados de la consulta
        print("Resultados de la consulta:", resultados)

        if resultados:
            # Paso 4: Creación del DataFrame
            datos = {"Cliente": [resultado[0] + " " + resultado[1] for resultado in resultados],
                     "Fecha de Pago": [resultado[2] for resultado in resultados],
                     "Dinero Recaudado": [resultado[3] for resultado in resultados]}
            df = pd.DataFrame(datos)

            # Paso 5: Agregar columna de "Dinero Total Recaudado"
            dinero_total_recaudado = df["Dinero Recaudado"].sum()
            df.loc[0, "Dinero Total Recaudado"] = dinero_total_recaudado

            # Paso 6: Ventana de diálogo para seleccionar la ruta de guardado
            ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                         filetypes=[("Excel Files", "*.xlsx")])

            # Paso 7: Guardar el DataFrame en un archivo de Excel
            df.to_excel(ruta_guardado, index=False)

            # Paso 8: Modificar el formato del archivo Excel
            workbook = openpyxl.load_workbook(ruta_guardado)
            sheet = workbook.active

            # Agregar columna "Dinero recaudado por mes" a la izquierda de "Dinero Total Recaudado"
            sheet.insert_cols(4)  # Insertar la columna en la cuarta posición (antes de "Dinero Total Recaudado")
            sheet.cell(row=1, column=4).value = "Dinero recaudado por mes"  # Título de la columna
            sheet.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)  # Texto en negrita

            # Calcular la suma de "Dinero Recaudado" por mes
            month_sums = {}  # Diccionario para almacenar la suma por mes

            for row_num in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_num, column=2)  # Columna de Fecha de Pago (B)
                if cell.value:
                    payment_date = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                    month = payment_date.month

                    if month not in month_sums:
                        month_sums[month] = 0
                    month_sums[month] += sheet.cell(row=row_num, column=3).value  # Sumar el valor de "Dinero Recaudado"

            # Mostrar el resultado de la suma solo en la primera celda de cada mes
            for month, total_sum in month_sums.items():
                for row_num in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row_num, column=2)  # Columna de Fecha de Pago (B)
                    if cell.value:
                        payment_date = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                        if payment_date.month == month:
                            cell = sheet.cell(row=row_num, column=4)
                            cell.value = total_sum
                            cell.fill = openpyxl.styles.PatternFill(start_color="8FFE62", end_color="8FFE62",
                                                                    fill_type="solid")  # Color de fondo
                            cell.font = openpyxl.styles.Font(bold=True)  # Texto en negrita
                            cell.border = openpyxl.styles.Border(
                                left=openpyxl.styles.Side(style='thick'))  # Agregar borde izquierdo
                            break  # Salir del bucle después de establecer el valor en la primera celda del mes

            # Ajustar las columnas para acomodar la nueva columna "Dinero recaudado por mes"
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            # Agregar borde grueso alrededor de la primera fila (títulos)
            for col_letter in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1):
                for cell in col_letter:
                    cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thick'))

            # Agregar bordes oscuros entre columnas
            for col_num in range(1, sheet.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col_num)

                first_cell = sheet[col_letter + "1"]
                first_cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thick'))

                # Agregar borde izquierdo a cada celda de la columna
                for row_num in range(1, sheet.max_row + 1):
                    cell = sheet[col_letter + str(row_num)]
                    if cell.value:
                        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thick'))

                if col_num < sheet.max_column:
                    # Agregar borde derecho a la última celda de cada columna, excepto para la última columna
                    last_cell = sheet[col_letter + str(sheet.max_row)]
                    last_cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='thick'))

            # Para la última columna, agregar un borde derecho grueso
            last_col_letter = openpyxl.utils.get_column_letter(sheet.max_column)
            for row_num in range(1, sheet.max_row + 1):
                last_cell = sheet[last_col_letter + str(row_num)]
                last_cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thick'),
                                                          right=openpyxl.styles.Side(style='thick'))

            # Agregar borde alrededor de toda la primera fila
            border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thick'),
                                            bottom=openpyxl.styles.Side(style='thick'),
                                            left=openpyxl.styles.Side(style='thick'),
                                            right=openpyxl.styles.Side(style='thick'))
            for col_num in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.border = border

            # Pintar el fondo de la celda debajo de la columna "Dinero total recaudado" en amarillo
            sheet.cell(row=2, column=sheet.max_column).fill = openpyxl.styles.PatternFill(start_color="FFFF00",
                                                                                          end_color="FFFF00",
                                                                                          fill_type="solid")
            # Agregar formato en negrita
            sheet.cell(row=2, column=sheet.max_column).font = openpyxl.styles.Font(bold=True)

            # Ajustar el ancho de las columnas al contenido
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

            fecha_pago_column = sheet['B']

            current_month = None
            for i in range(2, len(fecha_pago_column) + 1):
                cell = fecha_pago_column[i - 1]
                payment_date = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                month = payment_date.month

                if current_month is None or month != current_month:
                    if current_month is not None:
                        for col_num in range(1, sheet.max_column + 1):
                            last_row_cell = sheet.cell(row=i - 1, column=col_num)
                            last_row_cell.border = openpyxl.styles.Border(
                                bottom=openpyxl.styles.Side(style='thick'))

                    for col_num in range(1, sheet.max_column + 1):
                        first_row_cell = sheet.cell(row=i, column=col_num)
                        first_row_cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thick'))

                    current_month = month

                for col_num in range(1, sheet.max_column + 1):
                    last_row_cell = sheet.cell(row=len(fecha_pago_column), column=col_num)
                    last_row_cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thick'))

            workbook.save(ruta_guardado)
            workbook.close()

            messagebox.showinfo("Exportación exitosa", f"Datos exportados a {ruta_guardado}")

        else:
            messagebox.showwarning("Advertencia", "No hay datos para exportar.")

        conn.close()

        # Hacer que la ventana vuelva a estar en primer plano
        self.window.focus_force()

    def actualizar_fecha_hora(self):
        # Obtener la fecha y hora actuales
        now = time.strftime("%d-%m-%Y %H:%M:%S")

        # Actualizar la etiqueta con la fecha y hora actuales
        self.etiqueta_fecha_hora.config(text=now)

        # Verificar si se ha seleccionado una actividad para habilitar/deshabilitar el botón Exportar
        if self.actividad_seleccionada.get() != "--SELECCIONAR ACTIVIDAD--":
            self.boton_exportar["state"] = "normal"
        else:
            self.boton_exportar["state"] = "disabled"

        # Llamar a la función nuevamente después de 1000 ms (1 segundo)
        self.window.after(1000, self.actualizar_fecha_hora)

    def volver_pantalla_principal(self):
        # Cerrar la ventana de exportación solo si se decide volver a la pantalla principal
        self.window.destroy()


if __name__ == "__main__":
    ventana_exportar = tk.Tk()
    app_exportar = VentanaExportar(ventana_exportar)
    ventana_exportar.mainloop()




