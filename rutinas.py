import tkinter as tk
from tkinter import Label, Button, OptionMenu, Listbox, StringVar, messagebox, Checkbutton, simpledialog, Scrollbar, Frame
from tkinter.ttk import Combobox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import time
import sqlite3

class VentanaRutinas:
    def __init__(self, ventana_principal):
        self.ventana_principal = ventana_principal
        self.ventana_rutinas = tk.Toplevel(ventana_principal)
        self.ventana_rutinas.title("Creación de Rutina")
        self.ventana_rutinas.geometry("980x720")

        # Conexión a la base de datos SQLite
        self.conexion = sqlite3.connect("gimnasio.db")
        self.cursor = self.conexion.cursor()

        # Obtener las dimensiones de la pantalla
        screen_width = self.ventana_rutinas.winfo_screenwidth()
        screen_height = self.ventana_rutinas.winfo_screenheight()

        # Calcular las coordenadas para centrar la ventana
        x = (screen_width - 980) // 2
        y = (screen_height - 720) // 2

        # Establecer la geometría de la ventana
        self.ventana_rutinas.geometry(f"980x720+{x}+{y}")

        self.ventana_rutinas.resizable(width=False, height=False)
        self.ventana_rutinas.config(bg="#363636", bd=10)

        # Crear el marco principal
        marco_principal = Frame(self.ventana_rutinas, bg="#363636")
        marco_principal.pack(side="top", fill="x", pady=10, padx=10, anchor="nw")

        # Botón para regresar a la pantalla principal
        estilo_boton_ppal = {
            "bd": 0,
            "fg": "white",
            "relief": "flat",
            "font": ("Arial", 11),
            "padx": 10,
            "pady": 5,
            "bg": "#363636"
        }

        boton_principal = Button(
            marco_principal,
            text="Pantalla Principal",
            command=self.volver_pantalla_principal,
            **estilo_boton_ppal
        )
        boton_principal.pack(side="left")

        marco_fecha_hora = Frame(self.ventana_rutinas, bg="#212121", bd=2, relief="solid")
        marco_fecha_hora.place(relx=1, rely=1, anchor="se", x=-10, y=-10)

        # Etiqueta para la fecha y hora
        self.etiqueta_fecha_hora = Label(marco_fecha_hora, text="", font=("Arial", 12), bg="#212121", fg="#7FFFD4")
        self.etiqueta_fecha_hora.pack(padx=10, pady=5)

        # Actualizar la fecha y hora cada segundo
        self.actualizar_fecha_hora()

        # Agregar el título en negrita centrado sobre el marco de pagos
        titulo_label = Label(self.ventana_rutinas, text="CREAR RUTINA", font=("Arial", 24, "bold"), bg="#363636",
                             fg="#7FFFD4")
        titulo_label.pack(pady=(0, 10))

        self.grupos_musculares = ["Piernas", "Hombros", "Pecho", "Tríceps", "Bíceps", "Espalda", "Abdominales"]

        self.ejercicios_por_grupo = {
            "Piernas": ["Sentadillas", "Extensiones de Cuádriceps", "Prensa de Piernas", "Zancadas", "Estocadas",
                        "Sentadillas Búlgaras", "Gemelos en prensa", "Gemelos sentado", "Peso Muerto",
                        "Peso muerto con mancuernas"],
            "Hombros": ["Press Militar", "Elevaciones Laterales", "Pájaros", "Face Pulls", "Elevaciones frontales",
                        "Encogimiento", "Remo al mentón"],
            "Pecho": ["Press de Banca", "Pull-Over con Mancuerna", "Flexiones", "Aperturas con Cable",
                      "Plano con mancuernas", "Pecho inclinado", "Pecho declinado", "Flexiones diamante"],
            "Tríceps": ["Fondos en Paralelas", "Press Francés", "Rompecráneos", "Extensiones con Polea",
                        "Fondos en banco", "Patada de burro"],
            "Bíceps": ["Curl de Bíceps con Barra", "Curl Martillo", "Curl 21", "Curl de Concentración",
                       "Biceps en predicador", "Curl de Banca"],
            "Espalda": ["Dominadas", "Pull-Ups", "Remo con Barra T", "Pull-Downs", "Remo bajo", "Remo con barra"],
            "Abdominales": ['Crunches', 'Abdominal bolita', 'Abdominal vela', 'Elevación de piernas', 'Plancha',
                            'Plancha lateral', 'Crunches oblicuos', 'Bicicleta', 'Crunches con pelota',
                            'Rodillas al pecho', 'Crunches invertidos']
        }

        self.dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

        self.dia_seleccionado = StringVar()
        self.dia_seleccionado.set("")  # Configura el valor inicial como una cadena vacía

        # Combobox para seleccionar el día de la semana
        Label(self.ventana_rutinas, text="Selecciona un día de la semana:", font=("Arial", 12, "bold"), bg="#363636",
              fg="#7FFFD4").pack(pady=(10, 5))
        self.combo_dia_semana = Combobox(self.ventana_rutinas, values=self.dias_semana,
                                          textvariable=self.dia_seleccionado, state="readonly")
        self.combo_dia_semana.pack(pady=5)

        self.grupo_muscular_seleccionado = StringVar()
        self.grupo_muscular_seleccionado.set("")  # Configura el valor inicial como una cadena vacía

        # Combobox para seleccionar el grupo muscular
        Label(self.ventana_rutinas, text="Selecciona un grupo muscular:", font=("Arial", 12, "bold"), bg="#363636",
              fg="#7FFFD4").pack(pady=(10, 5))
        self.combo_grupos_musculares = Combobox(self.ventana_rutinas, values=self.grupos_musculares,
                                                textvariable=self.grupo_muscular_seleccionado, state="readonly")
        self.combo_grupos_musculares.pack(pady=5)

        # Frame para contener los botones
        frame_botones = tk.Frame(self.ventana_rutinas, bg="#363636")
        frame_botones.pack(pady=(10, 0))

        # Botones para cargar y agregar ejercicios
        cargar_ejercicios_btn = Button(frame_botones, text="Cargar Ejercicios", font=("Arial", 11, "bold"),
                                        bg="#7FFFD4",
                                        fg="#000000", command=self.cargar_ejercicios)
        cargar_ejercicios_btn.pack(side=tk.LEFT, padx=(5, 5))

        agregar_ejercicio_btn = Button(frame_botones, text="Agregar Nuevo Ejercicio", font=("Arial", 11, "bold"),
                                        bg="#7FFFD4",
                                        fg="#000000", command=self.agregar_ejercicio)
        agregar_ejercicio_btn.pack(side=tk.LEFT, padx=(5, 5))

        agregar_a_rutina_btn = Button(frame_botones, text="Agregar a Rutina", font=("Arial", 11, "bold"),
                                      bg="#7FFFD4",
                                      fg="#000000", command=self.agregar_a_rutina)
        agregar_a_rutina_btn.pack(side=tk.LEFT, padx=(5, 5))

        # Lista para almacenar ejercicios seleccionados por día
        self.ejercicios_por_dia = {}

        # Texto para mostrar ejercicios con scrollbar
        frame_texto = tk.Frame(self.ventana_rutinas)
        frame_texto.pack(pady=(10, 5))

        # Crear el widget de Texto una vez al inicio
        self.text_ejercicios = tk.Text(frame_texto, height=8, width=40)
        self.text_ejercicios.pack(side=tk.LEFT)

        scrollbar_texto = Scrollbar(frame_texto, command=self.text_ejercicios.yview)
        scrollbar_texto.pack(side=tk.RIGHT, fill=tk.Y)
        self.text_ejercicios.config(yscrollcommand=scrollbar_texto.set)

        # Deshabilitar la edición del Text
        self.text_ejercicios.bind("<Key>", lambda e: "break")

        # Frame para contener el Listbox y el Scrollbar
        frame_listbox = tk.Frame(self.ventana_rutinas)
        frame_listbox.pack(pady=(10, 5), padx=(5, 5))

        # Listbox para mostrar ejercicios seleccionados
        self.lista_ejercicios_seleccionados = Listbox(frame_listbox, selectmode=tk.MULTIPLE, height=8, width=40)
        self.lista_ejercicios_seleccionados.pack(side=tk.LEFT, fill=tk.BOTH)

        # Scrollbar para el Listbox
        scrollbar_listbox = Scrollbar(frame_listbox, orient=tk.VERTICAL,
                                      command=self.lista_ejercicios_seleccionados.yview)
        scrollbar_listbox.pack(side=tk.RIGHT, fill=tk.Y)

        # Nuevo Frame para contener el Botón
        frame_botones = tk.Frame(self.ventana_rutinas)
        frame_botones.pack(pady=(5, 5))

        # Botón para crear la rutina y exportar a Excel
        Button(frame_botones, text="Crear Rutina y Exportar a Excel", font=("Arial", 11, "bold"), bg="#7FFFD4",
               fg="#000000", command=self.crear_rutina_y_exportar_excel).pack()

    def cargar_ejercicios(self):
        grupo_muscular = self.grupo_muscular_seleccionado.get()
        ejercicios_grupo = self.obtener_ejercicios_por_grupo(grupo_muscular)

        self.text_ejercicios.config(state=tk.NORMAL)

        # Limpiar el contenido actual del widget de Texto
        self.text_ejercicios.delete(1.0, tk.END)

        # Asociar las variables de control directamente con los ejercicios
        self.check_vars = {ejercicio: tk.IntVar() for ejercicio in ejercicios_grupo}

        # Insertar los nuevos Checkbuttons en el widget de Texto
        for ejercicio in ejercicios_grupo:
            checkbox = Checkbutton(self.text_ejercicios, text=ejercicio, variable=self.check_vars[ejercicio])
            idx = self.text_ejercicios.window_create(tk.END, window=checkbox)
            self.text_ejercicios.insert(tk.END, "\n")

    def obtener_ejercicios_por_grupo(self, grupo_muscular):
        self.cursor.execute(
            "SELECT nombreEjercicio FROM Ejercicio WHERE idMusculo = (SELECT idMusculo FROM Musculo WHERE nombreMusculo = ?)",
            (grupo_muscular,))
        ejercicios = [row[0] for row in self.cursor.fetchall()]
        return ejercicios

    def agregar_ejercicio(self):
        grupo_muscular = self.grupo_muscular_seleccionado.get()
        ejercicio_nuevo = simpledialog.askstring("Agregar Nuevo Ejercicio",
                                                 "Ingresa el nombre del nuevo ejercicio para {}".format(
                                                     grupo_muscular))

        if ejercicio_nuevo:
            # Obtener el idMusculo correspondiente al grupo muscular seleccionado
            self.cursor.execute("SELECT idMusculo FROM Musculo WHERE nombreMusculo = ?", (grupo_muscular,))
            id_musculo = self.cursor.fetchone()[0]

            # Insertar el nuevo ejercicio en la base de datos
            self.cursor.execute("INSERT INTO Ejercicio (nombreEjercicio, idMusculo) VALUES (?, ?)",
                                (ejercicio_nuevo, id_musculo))
            self.conexion.commit()

            self.ejercicios_por_grupo.setdefault(grupo_muscular, []).append(ejercicio_nuevo)
            messagebox.showinfo("Éxito", "Ejercicio '{}' agregado correctamente.".format(ejercicio_nuevo))
        else:
            messagebox.showwarning("Advertencia", "Debes ingresar un nombre para el nuevo ejercicio.")

    def agregar_a_rutina(self):
        dia_seleccionado = self.dia_seleccionado.get()
        ejercicios_seleccionados = [
            (self.grupo_muscular_seleccionado.get(), ejercicio)
            for ejercicio, var in self.check_vars.items()
            if var.get()
        ]

        if not ejercicios_seleccionados:
            messagebox.showwarning("Advertencia", "Selecciona al menos un ejercicio para agregar a la rutina.")
            return

        # Agregar ejercicios al diccionario de ejercicios por día
        if dia_seleccionado not in self.ejercicios_por_dia:
            self.ejercicios_por_dia[dia_seleccionado] = []
        self.ejercicios_por_dia[dia_seleccionado].extend(ejercicios_seleccionados)

        # Actualizar el Listbox con los ejercicios seleccionados
        self.lista_ejercicios_seleccionados.delete(0, tk.END)
        for dia, ejercicios in self.ejercicios_por_dia.items():
            self.lista_ejercicios_seleccionados.insert(tk.END, f"{dia}:")
            for grupo, ejercicio in ejercicios:
                self.lista_ejercicios_seleccionados.insert(tk.END, f"   {grupo}: {ejercicio}")

    def crear_rutina_y_exportar_excel(self):

        # Verificar si hay ejercicios para exportar
        if not self.ejercicios_por_dia:
            messagebox.showwarning("Advertencia", "No hay ejercicios para exportar.")
            return

        # Crear un nuevo libro de trabajo de Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Rutina"

        # Estilo para títulos en negrita y borde
        title_style = Font(bold=True)
        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        # Escribir los datos en el libro de trabajo de Excel
        worksheet.append(["Día", "Grupo Muscular", "Ejercicio"])

        # Aplicar estilos a la primera fila (títulos)
        for cell in worksheet[1]:
            cell.font = title_style
            cell.border = border_style

        # Aplicar borde negro a la primera celda de cada columna
        for col in worksheet.columns:
            col[0].border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),
                                   bottom=Side(style='thick'))

        for dia, ejercicios in self.ejercicios_por_dia.items():
            for grupo, ejercicio in ejercicios:
                worksheet.append([dia, grupo, ejercicio])

        # Ajustar automáticamente el ancho de las columnas
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Agregar bordes a las celdas
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border_style

        # Guardar el libro de trabajo de Excel
        ruta_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                    filetypes=[("Archivos de Excel", "*.xlsx")])
        if ruta_archivo:
            workbook.save(ruta_archivo)
            messagebox.showinfo("Éxito", f"Rutina exportada a {ruta_archivo}")


        self.ventana_rutinas.deiconify()
        self.ventana_rutinas.focus_force()

    def actualizar_fecha_hora(self):
        # Obtener la fecha y hora actuales
        now = time.strftime("%d-%m-%Y %H:%M:%S")

        # Actualizar la etiqueta con la fecha y hora actuales
        self.etiqueta_fecha_hora.config(text=now)

        # Llamar a la función nuevamente después de 1000 ms (1 segundo)
        self.ventana_rutinas.after(1000, self.actualizar_fecha_hora)

    def volver_pantalla_principal(self):

        # Cerrar la ventana d solo si se decide volver a la pantalla principal
        self.ventana_rutinas.destroy()
        # Mostrar la pantalla principal nuevamente
        self.ventana_principal.deiconify()